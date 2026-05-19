import frappe
import random
from frappe.utils import getdate


def ensure_employment_types():
    types_to_create = [
        {"name": "CDI", "have_bonus": 1, "air_ticket": 1},
        {"name": "CDD", "have_bonus": 0, "air_ticket": 0},
    ]
    for t in types_to_create:
        if not frappe.db.exists("Employment Type", t["name"]):
            doc = frappe.new_doc("Employment Type")
            doc.name = t["name"]
            doc.employee_type_name = t["name"]
            doc.have_bonus = t.get("have_bonus", 0)
            doc.air_ticket = t.get("air_ticket", 0)
            doc.save(ignore_permissions=True, ignore_validate=True)
            print(f"Created Employment Type: {t['name']}")


def parse_name(full_name):
    parts = full_name.strip().split()
    if len(parts) >= 2:
        first_name = parts[-1]
        last_name = " ".join(parts[:-1])
    else:
        first_name = full_name
        last_name = ""
    return first_name, last_name


def estimate_dob(join_date):
    join = getdate(join_date) if join_date else getdate("2015-01-01")
    years_service = (getdate("2025-01-01").year - join.year)
    estimated_age = 25 + max(0, years_service)
    birth_year = join.year - estimated_age
    return getdate(f"{birth_year}-01-15")


def get_gender(name):
    male_names = {"CHRISTIAN", "GAUTHIER", "DIDIER", "DESCARTES", "LARRY", "CHRISTOPHE",
                  "DADDY", "SERGE", "PLATINI", "GHISLAIN", "BRUNO", "FULGENCE", "BLAISE",
                  "VICTOR", "GIRESSE", "ALAIN", "CEDRICK", "ERIC", "PETER", "STEVE",
                  "JOSEPH", "PATCHELY", "MALE", "ARISTOTE", "EVARISTE", "MOKE",
                  "MOHAMED", "ANTOINE", "PAULIN", "THIERRY", "BENOIT", "JEAN", "PIERRE",
                  "HENRI", "PHILIPPE", "JACQUES", "FRANCOIS", "ANDRE", "ROBERT", "LOUIS",
                  "MICHEL", "MARCEL", "ALBERT", "LEON", "EMILE", "PAUL", "DANIEL",
                  "GERARD", "MAURICE", "GEORGES", "CHARLES", "NICOLAS", "JULIEN",
                  "GUILLAUME", "ALEXANDRE", "MAXIME", "QUENTIN", "ANTHONY", "KEVIN",
                  "FLORIAN", "RAPHAEL", "BENJAMIN", "ADRIEN", "LUCAS", "KEVIN",
                  "THOMAS", "JONATHAN", "DAVID", "MATHIEU", "REMI", "YANNICK",
                  "CEDRIC", "OLIVIER", "XAVIER", "BRUNO", "DENIS", "SEBASTIEN",
                  "CHRISTOPHER", "FREDERIC", "VINCENT", "STEPHANE", "LAURENT",
                  "PASCAL", "ERIC", "MARTIN", "GREGORY", "JEREMY", "DAMIEN",
                  "SYLVAIN", "JORDAN", "AXEL", "MEHDI", "KARIM", "SAMIR", "HASSAN",
                  "YOUSSEF", "OMAR", "AHMED", "ABDEL", "MUSTAPHA", "RACHID",
                  "NASSER", "TARIK", "KHALID", "BRAHIM", "HAMZA", "IMAD", "ANAS",
                  "REDDA", "DJAMEL", "MOUNIR", "SOFIANE", "WALID", "ADEL",
                  "MOHAND", "MOURAD", "SALIM", "FAYCAL", "ILYAS", "YASSINE",
                  "AMINE", "OTMANE", "BADR", "NABIL", "ZAKARIA", "ISMAIL",
                  "YOUNES", "SAID", "ABDELKADER", "MOHAMMED", "ABDELLATIF",
                  "AZIZ", "DRIS", "MED", "SIMO", "FARID", "HICHAM", "JAMAL",
                  "LHOUCINE", "ABDELILAH", "RIDA", "TAOUFIK", "MUSTAFA",
                  "HAMID", "NOUREDDINE", "RACHID", "MOULAY", "HASSAN"}
    female_names = {"STEPHANIE", "COUCOU", "ANNIE", "RUTH", "SANDRA", "MAMY", "MIRANDA",
                    "FATUMA", "MIMIE", "BLANCHE", "VICTORINE", "MARIE", "JEANNE",
                    "MARGUERITE", "LUCIE", "THERESE", "MADELEINE", "YVONNE", "SIMONE",
                    "JULIETTE", "PAULE", "SUZANNE", "COLETTE", "FRANCOISE", "DENISE",
                    "NICOLE", "MONIQUE", "CHRISTIANE", "ANDREE", "CLAUDE", "JACQUELINE",
                    "GERMAINE", "ROSE", "JOSEPHINE", "CAMILLE", "EMMA", "MANON",
                    "JUSTINE", "MARION", "LAURA", "AUDREY", "MELANIE", "JULIE",
                    "SARAH", "ELISE", "CLEMENCE", "ALICE", "LOLA", "INÈS",
                    "EVA", "LEA", "CHLOE", "AMBRE", "MAELYS", "JADE", "LINA",
                    "ZOE", "LILOU", "ROSA", "LUCIA", "MIA", "SOPHIA", "AMELIA",
                    "VALENTINA", "MARTINA", "ELENA", "GABRIELA", "PATRICIA",
                    "SANDRINE", "NATHALIE", "VERONIQUE", "ISABELLE", "CATHERINE",
                    "ANNE", "CORINNE", "BRIGITTE", "SYLVIE", "CHANTAL", "DOMINIQUE",
                    "FLORENCE", "AGNES", "HELENE", "CECILE", "MARIELLE", "LAURENCE",
                    "CHRISTINE", "PERRINE", "ODILE", "SOLANGE", "GENEVIEVE",
                    "FRANCOISE", "JOSIANE", "LILIANE", "GISELE", "ROSEMARIE",
                    "Yvette", "Colette"}
    first = name.split()[-1].upper() if name else ""
    if first in male_names:
        return "Male"
    if first in female_names:
        return "Female"
    return random.choice(["Male", "Female"])


def import_from_excel():
    import openpyxl

    ensure_employment_types()
    frappe.db.commit()

    wb = openpyxl.load_workbook(
        "/home/frappe-user/frappe-bench/apps/be_pay/Documents/Provision/Vacation+13e ERP.xlsx",
        data_only=True,
    )
    ws = wb["VAC+13th (2)"]

    company = frappe.defaults.get_user_default("Company") or "Ointech Groupe"
    all_types = ["CDI", "CDD", "Contract", "Commission", "Apprentice"]

    # Pré-créer toutes les designations
    positions = set()
    for row in range(8, 300):
        emp_id = ws.cell(row=row, column=1).value
        if not emp_id:
            continue
        pos = str(ws.cell(row=row, column=3).value or "").strip()
        if pos:
            positions.add(pos)

    for pos in positions:
        if not frappe.db.exists("Designation", pos):
            des = frappe.new_doc("Designation")
            des.name = des.designation_name = pos
            des.save(ignore_permissions=True, ignore_validate=True)

    frappe.db.commit()
    print(f"Pre-created {len(positions)} designations")

    created = 0
    updated = 0
    errors = []
    rows_data = []

    for row in range(8, 300):
        emp_id = ws.cell(row=row, column=1).value
        full_name = ws.cell(row=row, column=2).value
        if not emp_id or not full_name:
            continue

        position = str(ws.cell(row=row, column=3).value or "").strip()
        salary = ws.cell(row=row, column=5).value
        join_date = ws.cell(row=row, column=7).value
        change_date1 = ws.cell(row=row, column=11).value
        new_base1 = ws.cell(row=row, column=12).value
        change_date2 = ws.cell(row=row, column=13).value
        new_base2 = ws.cell(row=row, column=14).value
        leaving_date = ws.cell(row=row, column=15).value
        contract_type = str(ws.cell(row=row, column=9).value or "").strip()

        first_name, last_name = parse_name(str(full_name))
        gender = get_gender(str(full_name))
        dob = estimate_dob(join_date)
        join = getdate(join_date) if join_date else getdate("2020-01-01")
        status = "Active" if not leaving_date else "Inactive"

        if contract_type in ("CDI", "CDD"):
            emp_type = contract_type
        else:
            emp_type = random.choice(all_types)

        if random.random() < 0.15:
            emp_type = random.choice(all_types)

        rows_data.append({
            "first_name": first_name,
            "last_name": last_name,
            "employee_name": f"{first_name} {last_name}".strip(),
            "gender": gender,
            "date_of_birth": dob,
            "date_of_joining": join,
            "relieving_date": getdate(leaving_date) if leaving_date else None,
            "status": status,
            "company": company,
            "employment_type": emp_type,
            "designation": position or None,
            "ctc": float(salary) if salary else 0,
            "housing": 0,
            "transport": 0,
            "allowance": 0,
            "change_date1": change_date1,
            "new_base1": new_base1,
            "change_date2": change_date2,
            "new_base2": new_base2,
        })

    # Création rapide
    for i, data in enumerate(rows_data):
        try:
            existing = frappe.db.get_value(
                "Employee", {"employee_name": data["employee_name"]}, "name"
            )

            if existing:
                doc = frappe.get_doc("Employee", existing)
                for k, v in data.items():
                    if k not in ("change_date1", "new_base1", "change_date2", "new_base2"):
                        setattr(doc, k, v)
                doc.flags.ignore_validate = True
                doc.save(ignore_permissions=True)
                updated += 1
            else:
                doc = frappe.new_doc("Employee")
                doc.naming_series = "HR-EMP-"
                for k, v in data.items():
                    if k not in ("change_date1", "new_base1", "change_date2", "new_base2"):
                        setattr(doc, k, v)
                doc.flags.ignore_validate = True
                doc.insert(ignore_permissions=True)
                created += 1

            # Pay Salary Storie
            stories = []
            if data["change_date1"] and data["new_base1"]:
                stories.append({
                    "salary": float(data["new_base1"]),
                    "start_date": getdate(data["change_date1"]),
                    "end_date": getdate(data["change_date2"]) if data["change_date2"] else None,
                })
            if data["change_date2"] and data["new_base2"]:
                stories.append({
                    "salary": float(data["new_base2"]),
                    "start_date": getdate(data["change_date2"]),
                    "end_date": None,
                })

            if stories:
                doc = frappe.get_doc("Employee", doc.name)
                doc.pay_salary_storie = []
                for s in stories:
                    doc.append("pay_salary_storie", s)
                doc.flags.ignore_validate = True
                doc.save(ignore_permissions=True)

            if (i + 1) % 50 == 0:
                frappe.db.commit()
                print(f"  ... processed {i+1}/{len(rows_data)}")

        except Exception as e:
            errors.append((data["employee_name"], str(e)))

    frappe.db.commit()
    print(f"DONE - Created: {created}, Updated: {updated}, Errors: {len(errors)}")
    if errors:
        for e in errors[:20]:
            print("  ERR:", e)


def run():
    frappe.flags.in_import = True
    import_from_excel()
    frappe.flags.in_import = False
