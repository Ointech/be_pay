#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script complet de création des DocTypes, controllers, hooks, JS, rapports
et print formats pour l'application be_pay à partir des fichiers Excel.
"""

import os
import sys
import json
import textwrap
import shutil
import re
import ast

# Assurer que openpyxl est dispo
try:
    import openpyxl
except ImportError:
    print("openpyxl non installé")
    sys.exit(1)

BASE_APP = "/home/frappe-user/frappe-bench/apps/be_pay"
BE_PAY_MOD = os.path.join(BASE_APP, "be_pay", "be_pay")
PUBLIC_JS = os.path.join(BASE_APP, "be_pay", "public", "js")
HOOKS_PY = os.path.join(BASE_APP, "be_pay", "hooks.py")
DOC_FOLDER = os.path.join(BASE_APP, "Documents")

sys.path.insert(0, "/home/frappe-user/frappe-bench/apps/frappe")
os.chdir("/home/frappe-user/frappe-bench")

# ---------------------------------------------------------------------------
# TRANSLATION MAPPINGS
# ---------------------------------------------------------------------------

DOCTYPE_NAME_MAP = {
    "Actualisation Categories": "Category Update",
    "Anciennete": "Seniority",
    "Anciennete Details": "Seniority Detail",
    "Annees Provisions": "Provision Year",
    "Attendance Line": "Attendance Line",
    "Attendance list": "Attendance List",
    "Attendance Retrieval Server": "Attendance Retrieval Server",
    "Attendance Retrieval Settings": "Attendance Retrieval Settings",
    "Bonus Provision": "Bonus Provision",
    "Conge Pris": "Leave Taken",
    "Custom Paie Settings": "Payroll Settings",
    "Decompte Final": "Final Settlement",
    "Dependant": "Dependent",
    "Details Annees Decompte": "Settlement Year Detail",
    "Details Provision Decompte": "Provision Settlement Detail",
    "Details Updating Attendance": "Attendance Update Detail",
    "Element de Voyage": "Travel Element",
    "Element de Voyage Allocation": "Travel Element Allocation",
    "Element de Voyage Application": "Travel Element Application",
    "Element de Voyage Application Details": "Travel Element Application Detail",
    "Element de Voyage Details": "Travel Element Detail",
    "Employee Attendance Tool New": "Employee Attendance Tool",
    "Employee loan Application": "Employee Loan Application",
    "Employment Type List": "Employment Type List",
    "Init Pro Details": "Init Provision Detail",
    "Init Provision": "Init Provision",
    "Loan Employee details": "Loan Employee Detail",
    "Periode Decompte": "Settlement Period",
    "Preparation Provision Decompte": "Provision Settlement Preparation",
    "Provision": "Provision",
    "Provision Bonus": "Provision Bonus",
    "Provision Conge": "Provision Leave",
    "Provision Decompte": "Provision Settlement",
    "Provision Details": "Provision Detail",
    "Provision Gratification": "Provision Gratuity",
    "Provision Gratifie Details": "Provision Gratuity Detail",
    "Provision Init": "Provision Init",
    "Provision Initialisation": "Provision Initialization",
    "Provision Ratio": "Provision Ratio",
    "Provision Ticket": "Provision Ticket",
    "Salary Slip Pris": "Salary Slip Taken",
    "Salary Type": "Salary Type",
    "Categories": "Category",
    "Categories Actualisees": "Updated Category",
}

# reverse for table names
OLD_DOCTYPE_NAMES = list(DOCTYPE_NAME_MAP.keys())
NEW_DOCTYPE_NAMES = list(DOCTYPE_NAME_MAP.values())

def translate_doctype_names(text):
    # sort by length desc
    for old_name, new_name in sorted(DOCTYPE_NAME_MAP.items(), key=lambda x: len(x[0]), reverse=True):
        text = text.replace(old_name, new_name)
    return text

FIELDNAME_MAP_CUSTOM = {
    "taux_actuel": "pay_exchange_rate",
    "date_derniere_taux": "pay_last_rate_date",
    "categorie": "pay_category",
    "categorie_initiale": "pay_initial_category",
    "est_initiale": "pay_is_initial",
    "base": "pay_base",
    "transport": "pay_transport",
    "logement": "pay_housing",
    "allocation": "pay_allowance",
    "categories_actualisees": "pay_updated_categories",
    "status": "pay_status",
    "annee": "pay_year",
    "annee_debut": "pay_start_year",
    "annee_fin": "pay_end_year",
    "anciennete": "pay_seniority",
    "anciennete_details": "pay_seniority_details",
    "annees_anciennete": "pay_seniority_years",
    "automatique": "pay_automatic",
    "type_preavis": "pay_notice_type",
    "motif_depart": "pay_departure_reason",
    "departement": "pay_department",
    "salaire_de_base": "pay_base_salary",
    "employee_category": "pay_employee_category",
    "starting_date": "pay_starting_date",
    "date_fin_contrat": "pay_contract_end_date",
    "preavis": "pay_notice",
    "jour_preste": "pay_worked_days",
    "conge_sur_preavis": "pay_leave_on_notice",
    "conge_compensatoire": "pay_compensatory_leave",
    "conge_non_pris": "pay_untaken_leave",
    "gratification": "pay_gratuity",
    "total_jours": "pay_total_days",
    "date_debut": "pay_start_date",
    "date_fin": "pay_end_date",
    "date_embauche": "pay_hiring_date",
    "nombre_annees": "pay_number_of_years",
    "nombre_jours": "pay_number_of_days",
    "fin_contrat": "pay_contract_end",
    "categories": "pay_categories",
    "salaire": "pay_salary",
    "prime": "pay_bonus",
    "unpaid_days": "pay_unpaid_days",
    "jrs_prestés_et_récupe": "pay_worked_and_recovered_days",
    "congé_payés_non_pris": "pay_untaken_paid_leave",
    "nuits_pretées": "pay_loaned_nights",
    "jours_malade": "pay_sick_days",
    "jours_fériés": "pay_public_holidays",
    "jours_suspensions": "pay_suspension_days",
    "allocation_sc": "pay_suspension_allowance_pct",
    "nbre_enfants": "pay_number_of_children",
    "allocations_fam": "pay_family_allowance",
    "ancienneté_details": "pay_seniority_details_text",
    "preavis_days": "pay_notice_days",
    "conge_sur_préavis": "pay_leave_on_notice_2",
    "conge_compensatoire": "pay_compensatory_leave_2",
    "gratification13è_mois": "pay_13th_month_gratuity",
    "période_décompte": "pay_settlement_period",
    "anciennete_rate": "pay_seniority_rate",
    "conge_days_5_years": "pay_leave_days_5_years",
    "jour_ouvrable": "pay_working_day",
    "working_hours_per_day": "pay_working_hours_per_day",
    "background_workers": "pay_background_workers",
    "anciennete_en_annee": "pay_seniority_in_years",
    "multiple_salary_in_period": "pay_multiple_salary_in_period",
    "employee_checking": "pay_employee_checking",
    "bonus_in_separate_slip": "pay_bonus_in_separate_slip",
    "automatic_seniority": "pay_automatic_seniority",
    "full_enqueue": "pay_full_enqueue",
    "annees": "pay_years",
    "montant": "pay_amount",
    "provision": "pay_provision",
    "absence": "pay_absence",
    "sunday_hours": "pay_sunday_hours",
    "hours_30": "pay_hours_30",
    "hours_60": "pay_hours_60",
    "night_hours": "pay_night_hours",
    "enabled": "pay_enabled",
    "server_name": "pay_server_name",
    "staff": "pay_staff",
    "url_base": "pay_url_base",
    "token": "pay_token",
    "username": "pay_username",
    "password": "pay_password",
    "interval_value": "pay_interval_value",
    "interval_unit": "pay_interval_unit",
    "last_run": "pay_last_run",
    "servers": "pay_servers",
    "table_employe": "pay_employee_table",
    "element_name": "pay_element_name",
    "quantite_an": "pay_quantity_per_year",
    "is_carry_forward": "pay_is_carry_forward",
    "allow_negative": "pay_allow_negative",
    "allow_over_allocation": "pay_allow_over_allocation",
    "nom_complet": "pay_full_name",
    "date_application": "pay_application_date",
    "date_depart": "pay_departure_date",
    "date_arrivee": "pay_arrival_date",
    "voyoage_allocation": "pay_travel_allocation",
    "code": "pay_code",
    "nom": "pay_name",
    "disponible": "pay_available",
    "quantite": "pay_quantity",
    "reste": "pay_remaining",
    "id_allocation": "pay_allocation_id",
    "en_cours": "pay_in_progress",
    "utilise": "pay_used",
    "filters_section": "pay_filters_section",
    "select_employees_section": "pay_select_employees_section",
    "employees_html": "pay_employees_html",
    "set_attendance_details_section": "pay_set_attendance_details_section",
    "marked_attendance_section": "pay_marked_attendance_section",
    "marked_attendance_html": "pay_marked_attendance_html",
    "late_entry": "pay_late_entry",
    "early_exit": "pay_early_exit",
    "period_start": "pay_period_start",
    "period_end": "pay_period_end",
    "loan_start_date": "pay_loan_start_date",
    "is_quinzaine": "pay_is_quinzaine",
    "number_of_installments": "pay_number_of_installments",
    "loan_amount": "pay_loan_amount",
    "installment_amount": "pay_installment_amount",
    "basic": "pay_basic",
    "repayment_method": "pay_repayment_method",
    "description": "pay_description",
    "employee_details": "pay_employee_details",
    "basic_salary": "pay_basic_salary",
    "ratio": "pay_ratio",
    "conge": "pay_leave",
    "ticket": "pay_ticket",
    "gratification": "pay_gratuity_amount",
    "bonus": "pay_bonus_amount",
    "matricule": "pay_matricule",
    "pris_ratio": "pay_taken_ratio",
    "pris_conge": "pay_taken_leave",
    "pris_gratification": "pay_taken_gratuity",
    "pris_ticket": "pay_taken_ticket",
    "pris_bonus": "pay_taken_bonus",
    "pourcent": "pay_percentage",
    "annee_a": "pay_year_1",
    "annee_b": "pay_year_2",
    "annee_c": "pay_year_3",
    "annee_d": "pay_year_4",
    "annee_e": "pay_year_5",
    "annee_f": "pay_year_6",
    "annee_g": "pay_year_7",
    "annee_h": "pay_year_8",
    "annee_i": "pay_year_9",
    "annee_j": "pay_year_10",
    "annee_k": "pay_year_11",
    "annee_l": "pay_year_12",
    "annee_m": "pay_year_13",
    "annee_n": "pay_year_14",
    "annee_0": "pay_year_15",
    "periode_provision": "pay_provision_period",
    "tauxjour": "pay_daily_rate",
    "salaire_prestation": "pay_service_salary",
    "salaire_de_nuit": "pay_night_salary",
    "conge_paye": "pay_paid_leave",
    "jour_feries": "pay_holiday_days",
    "conge_pc": "pay_leave_pc",
    "conge_maladie": "pay_sick_leave",
    "gatification": "pay_gratuity_2",
    "brut_imposable": "pay_taxable_gross",
    "allocation_suspension_de_contrat": "pay_contract_suspension_allowance",
    "brut_a_payer": "pay_gross_payable",
    "inss": "pay_inss",
    "ipr": "pay_ipr",
    "net_a_payer": "pay_net_payable",
    "somme": "pay_sum",
    "mois": "pay_month",
    "periode_date_begin": "pay_period_date_begin",
    "periode_date_end": "pay_period_date_end",
    "date_join": "pay_date_join",
    "date_begin": "pay_date_begin",
    "date_end": "pay_date_end",
    "date_quit": "pay_date_quit",
    "new_rate": "pay_new_rate",
    "period_days": "pay_period_days",
    "rate": "pay_rate",
    "start_period_days": "pay_start_period_days",
    "years_difference": "pay_years_difference",
    "year_div_5": "pay_year_div_5",
    "report": "pay_report",
    "janvier": "pay_january",
    "fevrier": "pay_february",
    "mars": "pay_march",
    "avril": "pay_april",
    "mai": "pay_may",
    "juin": "pay_june",
    "juillet": "pay_july",
    "aout": "pay_august",
    "septembre": "pay_september",
    "octobre": "pay_october",
    "novembre": "pay_november",
    "decembre": "pay_december",
    "pris": "pay_taken",
    "total": "pay_total",
    "jour": "pay_day",
    "fraction": "pay_fraction",
    "salary_type": "pay_salary_type",
    "eventual": "pay_eventual",
    "is_main_salary": "pay_is_main_salary",
    "event_name": "pay_event_name",
    "type": "pay_type",
    "date_naissance": "pay_birth_date",
    "current_anciennete": "pay_current_seniority",
    "new_anciennete": "pay_new_seniority",
    "first_start": "pay_first_start",
    "second_start": "pay_second_start",
    "first_end": "pay_first_end",
    "second_end": "pay_second_end",
    "find": "pay_find",
    "a1": "pay_a1", "a2": "pay_a2", "a3": "pay_a3", "a4": "pay_a4", "a5": "pay_a5",
    "a6": "pay_a6", "a7": "pay_a7", "a8": "pay_a8", "a9": "pay_a9", "a10": "pay_a10",
    "n1": "pay_n1", "n2": "pay_n2", "n3": "pay_n3",
    "custom_sm": "pay_custom_sm",
    "custom_presence": "pay_custom_presence",
    "custom_mise_a_pied": "pay_custom_mise_a_pied",
}

def translate_fieldnames_custom(text):
    for old_fn, new_fn in sorted(FIELDNAME_MAP_CUSTOM.items(), key=lambda x: len(x[0]), reverse=True):
        pattern = r'\b' + re.escape(old_fn) + r'\b'
        text = re.sub(pattern, new_fn, text)
    return text


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def ensure_dir(path):
    os.makedirs(path, exist_ok=True)

def ensure_init(path):
    init = os.path.join(path, "__init__.py")
    if not os.path.exists(init):
        with open(init, "w", encoding="utf-8") as f:
            f.write("")

def read_excel(path):
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = rows[0]
    data = []
    for row in rows[1:]:
        d = {}
        for h, v in zip(headers, row):
            if h is not None:
                d[h] = v
        data.append(d)
    return data

def clean_script(script):
    if not script:
        return ""
    if not isinstance(script, str):
        return ""
    # Line endings: handle \r\n and standalone \r
    script = script.replace('\r\n', '\n')
    script = script.replace('\r', '\n')
    # Excel escape artifacts
    script = script.replace(r'\_', '_')
    script = script.replace(r'\*', '*')
    script = script.replace(r'\-', '-')
    # Fix comma-comments: line starts with optional spaces then comma-space or comma
    lines = script.split('\n')
    out = []
    for line in lines:
        stripped = line.lstrip(' ')
        if stripped.startswith(', '):
            line = line.replace(', ', '# ', 1)
        elif stripped.startswith(','):
            line = line.replace(',', '#', 1)
        out.append(line)
    return '\n'.join(out)

def is_disabled(row):
    for k in ('Desactivé', 'Disabled', 'disabled', 'Activé', 'Enabled'):
        v = row.get(k)
        if v is not None:
            if k in ('Desactivé', 'Disabled', 'disabled'):
                if v == 1 or str(v).lower() in ('1', 'true', 'yes'):
                    return True
            if k in ('Activé', 'Enabled'):
                if v == 0 or str(v).lower() in ('0', 'false', 'no'):
                    return True
    return False

def slugify(name):
    return name.lower().replace(" ", "_").replace("-", "_")

def pascal_case(name):
    return "".join(x.capitalize() for x in name.split(" ") if x)

def to_snake_case(name):
    return name.lower().replace(" ", "_").replace("-", "_")

def normalize_indent(script):
    """Strip common leading whitespace from all non-empty lines."""
    lines = script.split('\n')
    min_indent = None
    for line in lines:
        stripped = line.lstrip(' ')
        if stripped:
            indent = len(line) - len(stripped)
            if min_indent is None or indent < min_indent:
                min_indent = indent
    if min_indent is None or min_indent == 0:
        return script
    result = []
    for line in lines:
        stripped = line.lstrip(' ')
        if stripped:
            if len(line) >= min_indent:
                result.append(line[min_indent:])
            else:
                result.append(stripped)
        else:
            result.append('')
    return '\n'.join(result)

def replace_doc_with_self(script):
    """Replace 'doc' variable with 'self' for controller methods."""
    result = []
    in_string = None
    escape_next = False
    i = 0
    while i < len(script):
        ch = script[i]
        if escape_next:
            result.append(ch)
            escape_next = False
            i += 1
            continue
        if ch == '\\' and in_string is not None:
            result.append(ch)
            escape_next = True
            i += 1
            continue
        if in_string is None:
            if ch in ('"', "'"):
                in_string = ch
                result.append(ch)
            else:
                if script[i:i+3] == 'doc':
                    prev_ok = (i == 0) or (not script[i-1].isalnum() and script[i-1] != '_')
                    next_ok = (i+3 >= len(script)) or (not script[i+3].isalnum() and script[i+3] != '_')
                    if prev_ok and next_ok:
                        result.append('self')
                        i += 3
                        continue
                result.append(ch)
        else:
            if ch == in_string:
                in_string = None
            result.append(ch)
        i += 1
    return ''.join(result)


def indent_script(script, spaces):
    """Add leading spaces to every non-empty line."""
    lines = script.split('\n')
    result = []
    for line in lines:
        if line.strip():
            result.append(' ' * spaces + line)
        else:
            result.append('')
    return '\n'.join(result)


def is_valid_python(script):
    """Check if script is valid Python syntax."""
    try:
        ast.parse(script)
        return True
    except (SyntaxError, IndentationError):
        return False


def wrap_in_exec(script, is_controller=False):
    """Wrap a script string in exec() call for safe execution."""
    # Escape backslashes and quotes for JSON-style string
    safe = json.dumps(script)
    if is_controller:
        return f"        script = {safe}\n        exec(script, {{\"self\": self, \"frappe\": frappe, \"doc\": self}})"
    else:
        return f"    script = {safe}\n    exec(script, {{\"doc\": doc, \"frappe\": frappe, \"self\": doc}})"


# ---------------------------------------------------------------------------
# 1. DOC TYPES
# ---------------------------------------------------------------------------

def create_doctype_json(name, module, issingle=0, istable=0, fields=None, permissions=None):
    dt_dir = os.path.join(BE_PAY_MOD, "doctype", slugify(name))
    ensure_dir(dt_dir)
    ensure_init(dt_dir)

    if permissions is None:
        permissions = [
            {
                "create": 1 if not issingle else 0,
                "delete": 1 if not issingle else 0,
                "email": 1,
                "export": 1,
                "print": 1,
                "read": 1,
                "report": 1,
                "role": "System Manager",
                "share": 1,
                "write": 1
            }
        ]

    doc = {
        "actions": [],
        "allow_rename": 1,
        "creation": "2024-01-01 00:00:00.000000",
        "doctype": "DocType",
        "engine": "InnoDB",
        "field_order": [f.get("fieldname") for f in (fields or [])],
        "fields": fields or [],
        "modified": "2024-01-01 00:00:00.000000",
        "modified_by": "Administrator",
        "module": module,
        "name": name,
        "owner": "Administrator",
        "permissions": permissions,
        "sort_field": "modified",
        "sort_order": "DESC",
        "states": []
    }

    if issingle:
        doc["issingle"] = 1
    if istable:
        doc["istable"] = 1
        doc["editable_grid"] = 1

    json_path = os.path.join(dt_dir, f"{slugify(name)}.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(doc, f, indent=2, ensure_ascii=False)

    # Controller template
    py_path = os.path.join(dt_dir, f"{slugify(name)}.py")
    class_name = pascal_case(name)
    with open(py_path, "w", encoding="utf-8") as f:
        f.write("# Copyright (c) 2026, ebamadernis@gmail.com and contributors\n")
        f.write("# For license information, please see license.txt\n\n")
        f.write("import frappe\n")
        f.write("from frappe.model.document import Document\n\n")
        f.write(f"class {class_name}(Document):\n")
        f.write("    pass\n")

    # test file
    test_path = os.path.join(dt_dir, f"test_{slugify(name)}.py")
    if not os.path.exists(test_path):
        with open(test_path, "w", encoding="utf-8") as f:
            f.write("# Copyright (c) 2026, ebamadernis@gmail.com and contributors\n")
            f.write("# For license information, please see license.txt\n\n")
            f.write("import frappe\n")
            f.write("from frappe.tests.utils import FrappeTestCase\n\n")
            f.write(f"class Test{class_name}(FrappeTestCase):\n")
            f.write("    pass\n")

    return dt_dir


def build_doctypes():
    print("[1/7] Construction des DocTypes...")

    # --- Category (standard) ---
    create_doctype_json(
        "Category",
        "Be Pay",
        fields=[
            {
                "fieldname": "pay_category",
                "fieldtype": "Data",
                "label": "Category",
                "reqd": 1,
                "unique": 1
            }
        ]
    )

    # --- Updated Category (child table) ---
    create_doctype_json(
        "Updated Category",
        "Be Pay",
        istable=1,
        fields=[
            {"fieldname": "pay_category", "fieldtype": "Data", "label": "Category", "in_list_view": 1},
            {"fieldname": "pay_base", "fieldtype": "Float", "label": "Base", "in_list_view": 1},
            {"fieldname": "pay_transport", "fieldtype": "Float", "label": "Transport", "in_list_view": 1},
            {"fieldname": "pay_housing", "fieldtype": "Float", "label": "Housing", "in_list_view": 1},
            {"fieldname": "pay_allowance", "fieldtype": "Float", "label": "Allowance", "in_list_view": 1},
        ]
    )

    # --- Category Update (Single) ---
    create_doctype_json(
        "Category Update",
        "Be Pay",
        issingle=1,
        fields=[
            {"fieldname": "pay_exchange_rate", "fieldtype": "Float", "label": "Exchange Rate", "reqd": 1},
            {"fieldname": "column_break_gi3dn", "fieldtype": "Column Break"},
            {"fieldname": "pay_last_rate_date", "fieldtype": "Date", "label": "Last Rate Date"},
            {"fieldname": "categories_section", "fieldtype": "Section Break", "label": "Categories"},
            {"fieldname": "pay_category", "fieldtype": "Link", "label": "Category", "options": "Category"},
            {"fieldname": "pay_initial_category", "fieldtype": "Data", "label": "Initial Category", "read_only": 1},
            {"fieldname": "column_break_dooe1", "fieldtype": "Column Break"},
            {"fieldname": "pay_status", "fieldtype": "Select", "label": "Status", "options": "\nIn Progress\nNew"},
            {"fieldname": "pay_is_initial", "fieldtype": "Check", "label": "Is Initial", "depends_on": "eval:doc.pay_status == 'New'"},
            {"fieldname": "section_break_uww0y", "fieldtype": "Section Break"},
            {"fieldname": "pay_base", "fieldtype": "Float", "label": "Base", "read_only": 1},
            {"fieldname": "column_break_rdgrk", "fieldtype": "Column Break"},
            {"fieldname": "pay_transport", "fieldtype": "Float", "label": "Transport", "read_only": 1},
            {"fieldname": "column_break_vjfuf", "fieldtype": "Column Break"},
            {"fieldname": "pay_housing", "fieldtype": "Float", "label": "Housing", "read_only": 1},
            {"fieldname": "column_break_duvf2", "fieldtype": "Column Break"},
            {"fieldname": "pay_allowance", "fieldtype": "Float", "label": "Allowance", "read_only": 1},
            {"fieldname": "pay_updated_categories_section", "fieldtype": "Section Break", "label": "Updated Categories"},
            {"fieldname": "pay_updated_categories", "fieldtype": "Table", "label": "Updated Categories", "options": "Updated Category"},
        ]
    )
    print("  -> 3 DocTypes créés")


# ---------------------------------------------------------------------------
# 2. SERVER SCRIPTS (Overrides)
# ---------------------------------------------------------------------------

def load_overrides():
    files = [
        os.path.join(DOC_FOLDER, "overrides", "Script de serveur (1).xlsx"),
        os.path.join(DOC_FOLDER, "overrides", "Script de serveur.xlsx"),
        os.path.join(DOC_FOLDER, "overrides", "Server Script.xlsx"),
    ]
    all_rows = []
    for f in files:
        if os.path.exists(f):
            all_rows.extend(read_excel(f))

    # Deduplicate by ID, keep first non-disabled
    seen = {}
    for row in all_rows:
        name = row.get("ID") or ""
        if not name:
            continue
        if name in seen:
            continue
        if is_disabled(row):
            seen[name] = None
            continue
        seen[name] = row

    return {k: v for k, v in seen.items() if v is not None}


def build_controllers_and_hooks(overrides):
    print("[2/7] Traitement des Server Scripts...")

    custom_doctype_events = {}
    standard_doctype_events = {}
    apis = []

    CUSTOM_DOCTYPES = {
        "Category Update",
        "Actualisation Categories",
        "Provision Settlement",
        "Provision Decompte",
        "Settlement Period",
        "Periode Decompte",
        "Provision Settlement Preparation",
        "Preparation Provision Decompte",
        "Final Settlement",
        "Decompte Final",
        "Employee Loan Application",
        "Employee loan Application",
        "Attendance List",
        "Attendance list",
        "Provision",
    }

    STANDARD_DOCTYPES = {
        "Employee",
        "Salary Slip",
        "Leave Application",
        "Employee Checkin",
        "Attendance",
        "Loan",
        "Leave Allocation",
        "Item",
    }

    for name, row in overrides.items():
        script = clean_script(row.get("Script") or "")
        if not script.strip():
            continue

        dt = (row.get("Type du document de référence") or row.get("Reference Document Type") or "").strip()
        event = (row.get("Événement DocType") or row.get("DocType Event") or "").strip()
        script_type = (row.get("Type de Script") or row.get("Script Type") or "").strip()
        api_method = (row.get("Méthode API") or row.get("API Method") or "").strip()

        if script_type == "API" or api_method:
            apis.append({"name": name, "method": api_method or slugify(name), "script": script})
            continue

        if not dt or not event:
            print(f"    WARN: {name} sans doctype/event -> ignoré")
            continue

        if dt in CUSTOM_DOCTYPES:
            slug = slugify(dt)
            custom_doctype_events.setdefault(slug, {}).setdefault(event, []).append({"name": name, "script": script})
        elif dt in STANDARD_DOCTYPES:
            standard_doctype_events.setdefault(dt, {}).setdefault(event, []).append({"name": name, "script": script})
        else:
            if os.path.exists(os.path.join(BE_PAY_MOD, "doctype", slugify(dt))):
                custom_doctype_events.setdefault(slugify(dt), {}).setdefault(event, []).append({"name": name, "script": script})
            else:
                print(f"    WARN: {name} référence doctype inconnu '{dt}' -> ignoré")

    # ------------------------------------------------------------------
    # Write custom doctype controllers
    # ------------------------------------------------------------------
    for slug, events in custom_doctype_events.items():
        # slug may be French; find English equivalent via DOCTYPE_NAME_MAP reverse
        dt_name = slug.replace("_", " ").title()
        for old_name, new_name in DOCTYPE_NAME_MAP.items():
            if slugify(old_name) == slug:
                dt_name = new_name
                break
        en_slug = slugify(dt_name)
        py_path = os.path.join(BE_PAY_MOD, "doctype", en_slug, f"{en_slug}.py")
        class_name = pascal_case(dt_name)

        methods = []
        for event, scripts in events.items():
            method_name = event.lower().replace(" ", "_")
            body_lines = []
            for s in scripts:
                scr = normalize_indent(s["script"])
                scr = translate_doctype_names(scr)
                scr = translate_fieldnames_custom(scr)
                scr_controller = replace_doc_with_self(scr)
                if is_valid_python(scr_controller):
                    body_lines.append(f"        # --- from script: {s['name']} ---")
                    body_lines.append(indent_script(scr_controller, 8))
                    body_lines.append("")
                else:
                    # Fall back to exec with original script (doc preserved)
                    body_lines.append(f"        # --- from script: {s['name']} (exec fallback) ---")
                    body_lines.append(wrap_in_exec(scr, is_controller=True))
                    body_lines.append("")
                    print(f"    WARN: {s['name']} -> exec fallback (indentation corrompue)")
            body = "\n".join(body_lines)
            methods.append(f"    def {method_name}(self):\n{body}")

        methods_str = "\n\n".join(methods)
        content = f"""# Copyright (c) 2026, ebamadernis@gmail.com and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document

class {class_name}(Document):
{methods_str}
"""
        os.makedirs(os.path.dirname(py_path), exist_ok=True)
        with open(py_path, "w", encoding="utf-8") as f:
            f.write(content)
        print(f"    Controller mis à jour: {slug}")

    # ------------------------------------------------------------------
    # Write standard doctype event files
    # ------------------------------------------------------------------
    events_dir = os.path.join(BE_PAY_MOD, "events")
    ensure_dir(events_dir)
    ensure_init(events_dir)

    doc_events_config = {}
    for dt, events in standard_doctype_events.items():
        file_slug = slugify(DOCTYPE_NAME_MAP.get(dt, dt))
        py_path = os.path.join(events_dir, f"{file_slug}.py")

        methods = []
        for event, scripts in events.items():
            method_name = event.lower().replace(" ", "_")
            body_lines = []
            for s in scripts:
                scr = normalize_indent(s["script"])
                scr = translate_doctype_names(scr)
                # Do NOT translate fieldnames for standard doctypes to avoid breaking custom fields on Employee etc.
                if is_valid_python(scr):
                    body_lines.append(f"    # --- from script: {s['name']} ---")
                    body_lines.append(indent_script(scr, 4))
                    body_lines.append("")
                else:
                    body_lines.append(f"    # --- from script: {s['name']} (exec fallback) ---")
                    body_lines.append(wrap_in_exec(scr, is_controller=False))
                    body_lines.append("")
                    print(f"    WARN: {s['name']} -> exec fallback (indentation corrompue)")

            body = "\n".join(body_lines)
            methods.append(f"def {method_name}(doc, method=None):\n{body}")

            dt_en = DOCTYPE_NAME_MAP.get(dt, dt)
            if dt_en not in doc_events_config:
                doc_events_config[dt_en] = {}
            doc_events_config[dt_en][event] = f"be_pay.events.{file_slug}.{method_name}"

        content = "import frappe\n\n" + "\n\n".join(methods) + "\n"
        with open(py_path, "w", encoding="utf-8") as f:
            f.write(content)
        print(f"    Event file créé: {file_slug}.py")

    # ------------------------------------------------------------------
    # Write API file
    # ------------------------------------------------------------------
    if apis:
        api_path = os.path.join(BE_PAY_MOD, "api.py")
        ensure_init(BE_PAY_MOD)
        api_lines = ["import frappe", "from frappe import _"]
        whitelisted = []
        for api in apis:
            method_name = api["method"] or slugify(api["name"])
            whitelisted.append(method_name)
            scr = normalize_indent(clean_script(api["script"]))
            scr = translate_doctype_names(scr)
            api_lines.append(f"\n@frappe.whitelist()")
            api_lines.append(f"def {method_name}():")
            if is_valid_python(scr):
                for line in scr.split("\n"):
                    api_lines.append(f"    {line}")
            else:
                api_lines.append(wrap_in_exec(scr, is_controller=False).replace("    exec", "    exec"))
                print(f"    WARN: API {api['name']} -> exec fallback")
        with open(api_path, "w", encoding="utf-8") as f:
            f.write("\n".join(api_lines) + "\n")
        print(f"    API file créé: api.py ({len(apis)} endpoints)")
    else:
        whitelisted = []
        api_path = os.path.join(BE_PAY_MOD, "api.py")
        if os.path.exists(api_path):
            os.remove(api_path)

    return doc_events_config, whitelisted


# ---------------------------------------------------------------------------
# 3. CLIENT SCRIPTS (JS)
# ---------------------------------------------------------------------------

def load_client_scripts():
    files = [
        os.path.join(DOC_FOLDER, "js", "Client Script.xlsx"),
        os.path.join(DOC_FOLDER, "js", "Script client (1).xlsx"),
        os.path.join(DOC_FOLDER, "js", "Script client.xlsx"),
    ]
    all_rows = []
    for f in files:
        if os.path.exists(f):
            all_rows.extend(read_excel(f))

    seen = {}
    for row in all_rows:
        name = row.get("ID") or ""
        if not name:
            continue
        if name in seen:
            continue
        if is_disabled(row):
            seen[name] = None
            continue
        seen[name] = row

    return {k: v for k, v in seen.items() if v is not None}


def build_client_scripts(scripts):
    print("[3/7] Traitement des Client Scripts...")

    target_files = {
        "leave_application.js": [],
        "employee.js": [],
        "loan.js": [],
        "reinitialization.js": [],
        "category_update.js": [],
        "init_provision.js": [],
    }

    for name, row in scripts.items():
        dt = (row.get("DocType") or "").strip()
        script = clean_script(row.get("Script") or "")
        if not dt or not script.strip():
            continue

        if dt == "Leave Application":
            target_files["leave_application.js"].append((dt, script))
        elif dt == "Employee":
            target_files["employee.js"].append((dt, script))
        elif dt == "Loan":
            target_files["loan.js"].append((dt, script))
        elif dt == "Reinitialization":
            target_files["reinitialization.js"].append((dt, script))
        elif dt == "Actualisation Categories":
            target_files["category_update.js"].append((dt, script))
        elif dt == "Init Provision":
            target_files["init_provision.js"].append((dt, script))
        else:
            fname = f"{slugify(dt)}.js"
            target_files.setdefault(fname, []).append((dt, script))

    ensure_dir(PUBLIC_JS)
    doctype_js_map = {}

    for fname, entries in target_files.items():
        if not entries:
            continue
        fpath = os.path.join(PUBLIC_JS, fname)
        parts = []
        for dt, script in entries:
            dt_en = DOCTYPE_NAME_MAP.get(dt, dt)
            script_en = translate_doctype_names(script)
            parts.append(f"// --- Script for {dt_en} ---")
            parts.append(script_en)
            parts.append("")
        content = "\n".join(parts) + "\n"
        with open(fpath, "w", encoding="utf-8") as f:
            f.write(content)
        print(f"    JS créé: {fname}")
        doctype_js_map[DOCTYPE_NAME_MAP.get(entries[0][0], entries[0][0])] = f"public/js/{fname}"

    return doctype_js_map


# ---------------------------------------------------------------------------
# 4. RAPPORTS & PRINT FORMATS
# ---------------------------------------------------------------------------

def load_rapports():
    files = [
        os.path.join(DOC_FOLDER, "rapports", "Format d'Impression.xlsx"),
        os.path.join(DOC_FOLDER, "rapports", "Print Format.xlsx"),
        os.path.join(DOC_FOLDER, "rapports", "Rapport (1).xlsx"),
        os.path.join(DOC_FOLDER, "rapports", "Rapport.xlsx"),
        os.path.join(DOC_FOLDER, "rapports", "Report.xlsx"),
    ]
    all_rows = []
    for f in files:
        if os.path.exists(f):
            all_rows.extend(read_excel(f))
    return all_rows


def build_reports_and_prints(rows):
    print("[4/7] Traitement des Rapports et Print Formats...")

    seen_reports = {}
    seen_prints = {}
    for row in rows:
        name = (row.get("Nom du Rapport") or row.get("Report Name") or row.get("ID") or "").strip()
        if not name:
            continue
        ref_dt = (row.get("Doctype de Réf.") or row.get("Ref DocType") or row.get("DocType") or "").strip()
        is_std = row.get("Est Standard") or row.get("Is Standard") or "No"
        rtype = (row.get("Type de Rapport") or row.get("Report Type") or "").strip()
        disabled = row.get("Desactivé") or row.get("Disabled") or 0
        if disabled == 1:
            continue
        module = (row.get("Module") or "Be Pay").strip()
        pf_type = (row.get("Type de Format d'Impression") or row.get("Print Format Type") or "").strip()

        if rtype == "Script Report":
            if name not in seen_reports:
                seen_reports[name] = {
                    "name": name,
                    "ref_doctype": ref_dt,
                    "is_standard": "Yes" if is_std in (1, "Yes", "Oui") else "No",
                    "report_type": rtype,
                    "module": module,
                    "letter_head": row.get("En-Tête") or row.get("Letter Head") or "",
                    "add_total_row": 1 if (row.get("Ajouter une Ligne Totale") or row.get("Add Total Row")) in (1, "Yes", 1.0) else 0,
                }
        elif pf_type == "Jinja":
            if name not in seen_prints:
                seen_prints[name] = {
                    "name": name,
                    "doc_type": ref_dt,
                    "module": module,
                    "standard": "No",
                    "custom_format": 1,
                    "print_format_type": "Jinja",
                    "disabled": 0,
                    "default_print_language": row.get("Langue d'impression par défaut") or row.get("Default Print Language") or "",
                }

    report_names = ["Provision Settlement", "Monthly Attendance", "Security Staff Position"]
    for rname in report_names:
        if rname not in seen_reports:
            print(f"    WARN: Rapport '{rname}' non trouvé dans les Excel")
            continue
        info = seen_reports[rname]
        rslug = slugify(rname)
        rdir = os.path.join(BE_PAY_MOD, "report", rslug)
        ensure_dir(rdir)
        ensure_init(rdir)

        rjson = {
            "add_total_row": info["add_total_row"],
            "columns": [],
            "creation": "2024-01-01 00:00:00.000000",
            "disabled": 0,
            "docstatus": 0,
            "doctype": "Report",
            "idx": 0,
            "is_standard": info["is_standard"],
            "json": "{}",
            "letter_head": info["letter_head"],
            "modified": "2024-01-01 00:00:00.000000",
            "modified_by": "Administrator",
            "module": info["module"],
            "name": rname,
            "owner": "Administrator",
            "prepared_report": 0,
            "ref_doctype": DOCTYPE_NAME_MAP.get(info["ref_doctype"], info["ref_doctype"]),
            "report_name": rname,
            "report_type": "Script Report",
            "roles": [{"role": "System Manager"}]
        }
        with open(os.path.join(rdir, f"{rslug}.json"), "w", encoding="utf-8") as f:
            json.dump(rjson, f, indent=2, ensure_ascii=False)

        py_stub = f"""# Copyright (c) 2026, ebamadernis@gmail.com and contributors
# For license information, please see license.txt

import frappe

def execute(filters=None):
    columns = []
    data = []
    return columns, data
"""
        with open(os.path.join(rdir, f"{rslug}.py"), "w", encoding="utf-8") as f:
            f.write(py_stub)

        rname_en = DOCTYPE_NAME_MAP.get(rname, rname)
        js_stub = f"""frappe.query_reports["{rname_en}"] = {{
    filters: [
    ]
}};
"""
        with open(os.path.join(rdir, f"{rslug}.js"), "w", encoding="utf-8") as f:
            f.write(js_stub)

        print(f"    Script Report créé: {rname}")

    pf_names = ["PAYSLIP", "PAYSLIP CSP", "FINAL SETTLEMENT", "Payroll Unsubmitted", "testPayroll", "Payslip"]
    used_slugs = set()
    for pname in pf_names:
        if pname not in seen_prints:
            print(f"    WARN: Print Format '{pname}' non trouvé dans les Excel")
            continue
        info = seen_prints[pname]
        pslug = slugify(pname)
        # handle slug collisions
        original_pslug = pslug
        counter = 1
        while pslug in used_slugs:
            pslug = f"{original_pslug}_{counter}"
            counter += 1
        used_slugs.add(pslug)
        pdir = os.path.join(BE_PAY_MOD, "print_format", pslug)
        ensure_dir(pdir)
        ensure_init(pdir)

        pjson = {
            "absolute_value": 0,
            "align_labels_right": 0,
            "creation": "2024-01-01 00:00:00.000000",
            "css": "",
            "custom_format": 1,
            "default_print_language": info["default_print_language"],
            "disabled": 0,
            "doc_type": DOCTYPE_NAME_MAP.get(info["doc_type"], info["doc_type"]),
            "docstatus": 0,
            "doctype": "Print Format",
            "font": "Default",
            "format_data": "[{\"fieldname\": \"print_heading_template\", \"fieldtype\": \"Custom HTML\", \"label\": \"Print Heading Template\", \"options\": \"<div class=\\\"print-format-header\\\"></div>\"}]",
            "html": "<div>Print Format {{ doc.name }}</div>",
            "idx": 0,
            "line_breaks": 0,
            "margin_bottom": 15.0,
            "margin_left": 15.0,
            "margin_right": 15.0,
            "margin_top": 15.0,
            "modified": "2024-01-01 00:00:00.000000",
            "modified_by": "Administrator",
            "module": info["module"],
            "name": pname,
            "owner": "Administrator",
            "page_number": "Hide",
            "print_format_builder": 0,
            "print_format_builder_beta": 0,
            "print_format_type": "Jinja",
            "raw_printing": 0,
            "show_section_headings": 0,
            "standard": "No"
        }
        with open(os.path.join(pdir, f"{pslug}.json"), "w", encoding="utf-8") as f:
            json.dump(pjson, f, indent=2, ensure_ascii=False)
        print(f"    Print Format créé: {pname}")


# ---------------------------------------------------------------------------
# 5. HOOKS.PY
# ---------------------------------------------------------------------------

def update_hooks(doc_events_config, doctype_js_map, whitelisted_methods):
    print("[5/7] Mise à jour de hooks.py...")

    with open(HOOKS_PY, "r", encoding="utf-8") as f:
        content = f.read()

    # Remove previous auto-generated section
    start_marker = "# --- BEGIN AUTO-GENERATED ---"
    end_marker = "# --- END AUTO-GENERATED ---"
    while start_marker in content:
        idx_start = content.find(start_marker)
        idx_end = content.find(end_marker, idx_start)
        if idx_end == -1:
            content = content[:idx_start]
        else:
            content = content[:idx_start] + content[idx_end + len(end_marker):]

    doc_events_block = "doc_events = {\n"
    for dt, events in sorted(doc_events_config.items()):
        doc_events_block += f'    "{dt}": {{\n'
        for evt, method in sorted(events.items()):
            doc_events_block += f'        "{evt}": "{method}",\n'
        doc_events_block += "    },\n"
    doc_events_block += "}\n"

    doctype_js_block = "doctype_js = {\n"
    for dt, path in sorted(doctype_js_map.items()):
        doctype_js_block += f'    "{dt}": "{path}",\n'
    doctype_js_block += "}\n"

    override_block = "override_whitelisted_methods = {\n"
    for method in whitelisted_methods:
        override_block += f'    "be_pay.api.{method}": "be_pay.api.{method}",\n'
    override_block += "}\n"

    insert = f"\n{start_marker}\n"
    insert += f"# Auto-generated by create_all.py\n{doc_events_block}\n{doctype_js_block}\n"
    if whitelisted_methods:
        insert += f"{override_block}\n"
    insert += f"{end_marker}\n"

    content = content.rstrip() + "\n" + insert

    with open(HOOKS_PY, "w", encoding="utf-8") as f:
        f.write(content)
    print("  -> hooks.py mis à jour")


# ---------------------------------------------------------------------------
# 6. MIGRATION
# ---------------------------------------------------------------------------

def run_migration():
    print("[6/7] Exécution de la migration bench...")
    import subprocess
    result = subprocess.run(
        ["bench", "--site", "gestion.fr", "migrate"],
        cwd="/home/frappe-user/frappe-bench",
        capture_output=True,
        text=True
    )
    print(result.stdout)
    if result.returncode != 0:
        print("STDERR:", result.stderr)
        sys.exit(1)
    print("  -> Migration terminée")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print("CREATE_ALL.PY - Génération complète be_pay")
    print("=" * 60)

    build_doctypes()

    overrides = load_overrides()
    doc_events_config, whitelisted = build_controllers_and_hooks(overrides)

    client_scripts = load_client_scripts()
    doctype_js_map = build_client_scripts(client_scripts)

    rapports = load_rapports()
    build_reports_and_prints(rapports)

    update_hooks(doc_events_config, doctype_js_map, whitelisted)

    run_migration()

    print("=" * 60)
    print("TERMINE")
    print("=" * 60)


if __name__ == "__main__":
    main()
